"""
Excel 服务类
处理 Excel 文件的业务逻辑
"""

import io
import logging
from typing import List, BinaryIO
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

class ExcelService:
    """Excel 处理服务"""
    
    def get_columns_info(self, file_content: bytes) -> List[dict]:
        """
        获取 Excel 文件的列信息
        
        Args:
            file_content: Excel 文件的二进制内容
        
        Returns:
            List[dict]: 列信息列表
        """
        try:
            # 从字节流加载工作簿
            file_stream = io.BytesIO(file_content)
            workbook = load_workbook(file_stream, data_only=True)
            
            # 获取第一个工作表
            worksheet = workbook.active
            
            columns_info = []
            
            if worksheet.max_row > 0 and worksheet.max_column > 0:
                # 获取第一行作为表头
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    column_name = str(cell_value) if cell_value is not None else f"列{col}"
                    
                    # 获取该列的一些示例数据
                    sample_data = []
                    for row in range(2, min(6, worksheet.max_row + 1)):  # 最多取4行示例数据
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value is not None:
                            sample_data.append(str(cell_value))
                    
                    columns_info.append({
                        "index": col,
                        "name": column_name,
                        "sample_data": sample_data[:3]  # 只返回前3个示例
                    })
            
            logger.info(f"成功获取列信息，共 {len(columns_info)} 列")
            return columns_info
            
        except Exception as e:
            logger.error(f"获取列信息时出错: {str(e)}", exc_info=True)
            raise Exception(f"获取列信息失败: {str(e)}")
    
    def delete_columns(self, file_content: bytes, column_indices: List[int]) -> bytes:
        """
        删除 Excel 文件中的指定列
        
        Args:
            file_content: Excel 文件的二进制内容
            column_indices: 要删除的列索引列表（从1开始，降序排列）
        
        Returns:
            bytes: 处理后的 Excel 文件二进制内容
        
        Raises:
            Exception: 当处理过程中出现错误时
        """
        try:
            # 从字节流加载工作簿
            file_stream = io.BytesIO(file_content)
            workbook = load_workbook(file_stream, data_only=False)
            
            logger.info(f"成功加载工作簿，包含 {len(workbook.worksheets)} 个工作表")
            
            # 处理每个工作表
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                logger.info(f"处理工作表: {sheet_name}")
                
                # 检查工作表是否有数据
                if worksheet.max_row == 1 and worksheet.max_column == 1:
                    # 空工作表，跳过
                    continue
                
                # 验证列索引是否有效
                max_column = worksheet.max_column
                invalid_columns = [col for col in column_indices if col > max_column]
                if invalid_columns:
                    logger.warning(f"工作表 {sheet_name} 中的无效列索引: {invalid_columns} (最大列数: {max_column})")
                
                # 删除有效的列（从右到左删除，避免索引变化）
                valid_columns = [col for col in column_indices if col <= max_column]
                for col_index in valid_columns:
                    logger.info(f"删除工作表 {sheet_name} 的第 {col_index} 列")
                    self._delete_column_with_style(worksheet, col_index)
            
            # 保存到字节流
            output_stream = io.BytesIO()
            workbook.save(output_stream)
            output_stream.seek(0)
            
            result = output_stream.getvalue()
            logger.info(f"成功处理 Excel 文件，输出大小: {len(result)} 字节")
            
            return result
            
        except Exception as e:
            logger.error(f"处理 Excel 文件时出错: {str(e)}", exc_info=True)
            raise Exception(f"Excel 文件处理失败: {str(e)}")
    
    def _delete_column_with_style(self, worksheet: Worksheet, column_index: int):
        """
        删除指定列并尽量保持样式和合并单元格
        
        Args:
            worksheet: 工作表对象
            column_index: 要删除的列索引（从1开始）
        """
        try:
            # 获取要删除的列字母
            from openpyxl.utils import get_column_letter
            column_letter = get_column_letter(column_index)
            
            # 处理合并单元格
            # 需要先收集所有相关的合并单元格范围，然后再处理
            merged_ranges_to_remove = []
            merged_ranges_to_modify = []
            
            for merged_range in list(worksheet.merged_cells.ranges):
                min_col = merged_range.min_col
                max_col = merged_range.max_col
                
                if column_index >= min_col and column_index <= max_col:
                    # 合并范围包含要删除的列
                    if min_col == max_col == column_index:
                        # 合并范围只有这一列，删除整个合并
                        merged_ranges_to_remove.append(merged_range)
                    else:
                        # 合并范围跨多列，需要调整
                        merged_ranges_to_remove.append(merged_range)
                        if max_col > column_index:
                            # 创建新的合并范围
                            new_min_col = min_col if min_col < column_index else min_col - 1
                            new_max_col = max_col - 1
                            if new_max_col >= new_min_col:
                                from openpyxl.worksheet.cell_range import CellRange
                                new_range = CellRange(
                                    min_col=new_min_col,
                                    min_row=merged_range.min_row,
                                    max_col=new_max_col,
                                    max_row=merged_range.max_row
                                )
                                merged_ranges_to_modify.append(new_range)
                elif column_index < min_col:
                    # 要删除的列在合并范围左侧，需要调整合并范围的列索引
                    merged_ranges_to_remove.append(merged_range)
                    from openpyxl.worksheet.cell_range import CellRange
                    new_range = CellRange(
                        min_col=min_col - 1,
                        min_row=merged_range.min_row,
                        max_col=max_col - 1,
                        max_row=merged_range.max_row
                    )
                    merged_ranges_to_modify.append(new_range)
            
            # 删除旧的合并范围
            for merged_range in merged_ranges_to_remove:
                worksheet.unmerge_cells(str(merged_range))
            
            # 删除列
            worksheet.delete_cols(column_index)
            
            # 添加新的合并范围
            for new_range in merged_ranges_to_modify:
                try:
                    worksheet.merge_cells(str(new_range))
                except Exception as e:
                    logger.warning(f"重新合并单元格时出错: {str(e)}")
            
            logger.debug(f"成功删除第 {column_index} 列")
            
        except Exception as e:
            logger.error(f"删除列 {column_index} 时出错: {str(e)}")
            # 如果样式处理失败，尝试简单删除
            try:
                worksheet.delete_cols(column_index)
                logger.info(f"使用简单方式删除第 {column_index} 列")
            except Exception as e2:
                logger.error(f"简单删除列也失败: {str(e2)}")
                raise e2